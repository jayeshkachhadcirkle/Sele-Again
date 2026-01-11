import time
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl import worksheet
from plyer import notification


url = "https://www.google.com/maps/search/it+company+in+surat/"
# url = "https://www.google.com/maps/search/handicraft+in+gandhinagar/"

wb = load_workbook('FIN GNC.xlsx')
ws = wb.active

maxrow = ws.max_row
s_max = maxrow

driver = webdriver.Chrome()
driver.get(url)
driver.maximize_window()
time.sleep(10)

element = driver.find_element(By.XPATH, "//h1[normalize-space()='Results']")
# ActionChains(driver).scroll_to_element(element).perform()

scrollorigin = ScrollOrigin.from_element(element)
print(scrollorigin)
scrolltarget = 3000

e1 = 0

def grab_data():
    tiles = driver.find_elements(By.CSS_SELECTOR, "a[class='hfpxzc']")
    cons = driver.find_elements(By.XPATH, "//span[@class='UsdlK']")
    maplist = driver.find_elements(By.CSS_SELECTOR, "a[class='hfpxzc")

    for i in range(len(tiles)):
        name = tiles[i].get_attribute('aria-label')
        try:
            ph = cons[i].text
        except:
            ph = "null"
        href = maplist[i].get_attribute('href')

        print(name," = ", ph)
        ws.cell(row=i+1+maxrow, column=2).value = ph
        ws.cell(row=i+1+maxrow, column=3).value = href
        ws.cell(row=i+1+maxrow, column=1).value = name

while True:
    try:
        driver.find_element(By.XPATH, "//span[@class='HlvSq']")
        break
        print("E0.1")
        grab_data()
    except:
        try:
            ActionChains(driver).scroll_from_origin(scrollorigin, 0, scrolltarget).perform()
            # scrollorigin = scrollorigin + 3000
            scrolltarget = scrolltarget + 4000
            # print(scrolltarget)
            grab_data()
            time.sleep(3)
            e1 = 0
            wb.save('FIN GNC.xlsx')
        except:
            time.sleep(2)
            print("E1")
            e1 = e1 + 1
            if e1 > 10:
                # grab_data()
                # wb.save('FIN GNC.xlsx')
                break
            else:
                pass
            continue

time.sleep(2)

wb.save('FIN GNC.xlsx')
maxrow = ws.max_row
counts = maxrow - s_max

# grab_data()
if e1 > 10:
    notification.notify("MAP", "Error")
else:
    notification.notify("Succeeded with", str(counts))

ws.cell(row= maxrow + 1, column=1).value = "xxxxx" + url
wb.save('FIN GNC.xlsx')

# //span[@class='UsdlK']