import time
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

wb = load_workbook('FIN GNC.xlsx')
ws = wb.active


driver = webdriver.Chrome()
driver.get("https://www.google.com/maps/search/finance+in+ahmedabad/@23.0346499,72.4901484,12z?entry=ttu")
driver.maximize_window()
time.sleep(5)


tiles = driver.find_elements(By.CSS_SELECTOR, "a[class='hfpxzc']")
for i in range(len(tiles)):
    name = tiles[i].get_attribute('aria-label')
    # print(name)
    ws.cell(row=i+1, column=1).value = name

cons = driver.find_elements(By.XPATH, "//span[@class='UsdlK']")
for i in range(len(cons)):
    ph = cons[i].text
    # print(ph)
    ws.cell(row=i+1, column=2).value = ph

wb.save('FIN GNC.xlsx')

# for i in range(len(maplist)):
#     href = maplist[i].get_attribute('href')
#     print(href)


# //span[@class='UsdlK']