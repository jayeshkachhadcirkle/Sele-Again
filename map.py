import time
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl import worksheet

wb = load_workbook('FIN GNC.xlsx')
ws = wb.active

maxrow = ws.max_row

url = ""

driver = webdriver.Chrome()
driver.get(url)
driver.maximize_window()
time.sleep(5)

# time.sleep(2)

element = driver.find_element(By.XPATH, "//h1[normalize-space()='Results']")
# ActionChains(driver).scroll_to_element(element).perform()

scrollorigin = ScrollOrigin.from_element(element)
print(scrollorigin)
scrolltarget = 3000

def grab_data():
    tiles = driver.find_elements(By.CSS_SELECTOR, "a[class='hfpxzc']")
    for i in range(len(tiles)):
        name = tiles[i].get_attribute('aria-label')
        print(name)
        ws.cell(row=i+1+maxrow, column=1).value = name

    cons = driver.find_elements(By.XPATH, "//span[@class='UsdlK']")
    for i in range(len(cons)):
        ph = cons[i].text
        print(ph)
        ws.cell(row=i+1+maxrow, column=2).value = ph

    maplist = driver.find_elements(By.CSS_SELECTOR, "a[class='hfpxzc")
    for i in range(len(maplist)):
        href = maplist[i].get_attribute('href')
        ws.cell(row=i+1+maxrow, column=3).value = href
        # print(href)

while True:
    try:
        driver.find_element(By.XPATH, "//span[@class='HlvSq']")
        print("E0.1")
        grab_data()
        break
    except:
        try:
            ActionChains(driver).scroll_from_origin(scrollorigin, 0, scrolltarget).perform()
            # scrollorigin = scrollorigin + 3000
            scrolltarget = scrolltarget + 5000
            # print(scrolltarget)
            time.sleep(3)
        except:
            print("E1")
            continue

time.sleep(2)

grab_data()

wb.save('FIN GNC.xlsx')

# //span[@class='UsdlK']