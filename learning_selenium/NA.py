import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook

# wb = Workbook()
wb = load_workbook('NA_data.xlsx')
ws = wb.active

driver = webdriver.Chrome()
driver.get("https://goodporn.to/search/Naughty-America/")
driver.maximize_window()

time.sleep(2)

# pages = []
vidName = []
vidlinks = []
thums = []

try:
    for i in range(200):
        driver.find_element(By.XPATH, "(//li[@class='next'])[1]").click()
        time.sleep(2)
        vl = driver.find_elements(By.CLASS_NAME, "item")
        for l in range(24):
            n = vl[l].find_element(By.TAG_NAME, 'a').get_attribute('title')
        for k in range(24):
            ths = vl[k].find_element(By.CLASS_NAME, 'thumb').get_attribute('src')
            thums.append(ths)
        for j in range(24):
            href = vl[j].find_element(By.TAG_NAME, 'a').get_attribute('href')
            vidlinks.append(href)
        # vl = driver.find_elements(By.CSS_SELECTOR, "item > a")
        # print(vidlinks)
except:
    for i in vidlinks:
        print(i)
    #     ws.cell(row=i + 1, column=1).value = vidName[i]
    #     ws.cell(row=i + 1, column=2).value = thums[i]
    #     ws.cell(row=i + 1, column=3).value = vidlinks[i]
    #
    # wb.save('NA_data.xlsx')
    # print()
    print("Loop over")
