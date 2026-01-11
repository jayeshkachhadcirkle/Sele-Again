from faker import Faker
from openpyxl import Workbook

fake_data = Faker()
# fake_data = Faker('hi_IN') for hindi name or data
wb = Workbook()
ws = wb.active

print(fake_data.name())
print(fake_data.address())

for i in range(1,51):
    ws.cell(row=i, column=1).value = fake_data.name()
    ws.cell(row=i, column=2).value = fake_data.email()
    ws.cell(row=i, column=3).value = fake_data.address()
wb.save('fake_data.xlsx')

# for i in range(1,51):
#     ws.cell(row=i, column=2).value = fake_data
