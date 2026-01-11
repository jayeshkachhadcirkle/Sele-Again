import time
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl import worksheet

wb = load_workbook('FIN GNC.xlsx')
ws = wb.active

mx = ws.max_row
print(mx)

driver = webdriver.Chrome()

for i in range(321):
    driver.get(ws.cell(row = i+1, column = 3).value)
    time.sleep(1)
    try:
        try:
            name = driver.find_element(By.CSS_SELECTOR, ".DUwDvf.lfPIob").text
            ws.cell(row=i + 1, column=1).value = name
        except:
            ws.cell(row=i + 1, column=1).value = "errorrr"
            pass
        try:
            # ph = driver.find_element(By.XPATH, "(//div[@class='AeaXub'])[2]").text
            ph = driver.find_element(By.XPATH, "(//div[contains(@class,'rogA2c')])[2]").text
            ws.cell(row=i + 1, column=2).value = ph
            wb.save("FIN GNC.xlsx")
        except:
            ws.cell(row=i + 1, column=2).value = "errorrr"
            wb.save("FIN GNC.xlsx")
            pass
        print(ph)
    except:
        wb.save("FIN GNC.xlsx")
        continue
