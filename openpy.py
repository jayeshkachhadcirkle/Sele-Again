from openpyxl import load_workbook
from openpyxl import worksheet

wb = load_workbook("C:\\Users\\jayes\Desktop\contacts.xlsx")
ws = wb.active

mx = ws.max_row
print(mx)

# print(ws.cell(row=1, column=1).value)
newstr = "111111111111"

for i in range(mx):
    if len(str(ws.cell(row=i+1, column=1).value)) == 10:
        newcon = "91" + str(ws.cell(row=i+1, column=1).value)
        print(newcon)
    elif len(str(ws.cell(row=i+1, column=1).value)) == 12:
        print(str(ws.cell(row=i+1, column=1).value))
    else:
        pass