from openpyxl import load_workbook
from openpyxl import worksheet
import pandas as pd

print("Starting Compare")


# wb = load_workbook('inp50.xlsx')
# ws = wb.active
# maxrow = ws.max_row
# print("Max Row: ", maxrow)

pd.set_option('display.max_rows', 1000)
pd.set_option('display.max_columns', None)

# Sr	STYLE	SEASON	CATEGORY	COLOUR	SKU	SIZE	QTY	match

requiredpos = ['Sr', 'SIZE', 'COLOUR', 'STYLE', 'match', 'SKU']
dfpos = pd.read_excel('inp50.xlsx', sheet_name='inp50')
# count the number of rows in the dataframe
print("Total Rows in inp50.xlsx: ", len(dfpos))
dfposSelected = dfpos[requiredpos]
# print(dfposSelected.head(100))

required_cols = ['Sr', 'Option1 Value', 'Option2 Value', 'Variant SKU', 'match', 'resku']
dfad = pd.read_excel('iproscure.xlsx', sheet_name='proscure')
print("Total Rows in iproscure.xlsx: ", len(dfad))
dfadSelected = dfad[required_cols]

# Merge dataframes on 'match' column
merged_df = dfadSelected.merge(dfposSelected[['match', 'SKU']], on='match', how='left')

# Copy SKU from inp50.xlsx to 'resku' column
merged_df['resku'] = merged_df['SKU']

# Drop the temporary SKU column and keep resku
result_df = merged_df.drop(columns=['SKU'])

print("\nMerged Result:")
# print(result_df.head(100))

# Save to results.xlsx
result_df.to_excel('results.xlsx', sheet_name='results', index=False)
print("row count in results.xlsx: ", len(result_df))
print("\nResults saved to results.xlsx")