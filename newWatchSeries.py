import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

wb = load_workbook('save_data.xlsx')
ws = wb.active

driver = webdriver.Chrome()
driver.get("https://watchseries8.to/movies")
driver.maximize_window()

time.sleep(2)
thumbs_name = driver.find_elements(By.CSS_SELECTOR, "h3[class='movie-name']")
thumbs_link = driver.find_elements(By.CSS_SELECTOR, "a[class='movie-link")
# th_href = thumbs_link.get_attribute("href")

time.sleep(2)

print(len(thumbs_name))

for i in range(len(thumbs_name)):
    ws.cell(row=i+1, column=1).value = thumbs_name[i].text

for i in range(len(thumbs_link)):
    href = thumbs_link[i].get_attribute('href')
    ws.cell(row=i+1, column=2).value = href

wb.save('save_data.xlsx')

for i in range(24):
    driver.get(ws.cell(row=i+1, column=2).value)

    time.sleep(1)


