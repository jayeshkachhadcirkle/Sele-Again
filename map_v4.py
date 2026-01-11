import time
import random
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium_stealth import stealth
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl import worksheet
from plyer import notification

# url = "https://www.google.com/maps/search/handicraft+in+ahmedabad/"

# Rajasthan features vibrant cities like the capital historic Udaipur, the "City of Lakes"; the desert city of Jaisalmer, known for its golden fort; Jodhpur, the "Blue City" with its Mehrangarh Fort; and cultural hubs like Ajmer, Kota, and Bikaner, offering rich heritage, palaces, and unique desert landscapes.

# url = "https://www.google.com/maps/search/cafe+near+Jodhpur/"
url = "https://www.google.com/maps/search/cafe+near+noida/"
# url = "https://www.google.com/maps/search/cafe+near+Kota/"

wb = load_workbook('L_gen.xlsx')
ws = wb.active

maxrow = ws.max_row
s_max = maxrow

options = Options()

user_agents = [
    # Add your list of user agents here
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
    'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 13_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15',
]

user_agent = random.choice(user_agents)
options.add_argument("--disable-blink-features=AutomationControlled")
# options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument(f'user-agent={user_agent}')
options.add_argument('--disable-extensions')

driver = webdriver.Chrome(options=options)

stealth(driver,
        languages=["en-US", "en"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True)

driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

driver.get(url)
driver.maximize_window()
time.sleep(6)

element = driver.find_element(By.XPATH, "//h1[normalize-space()='Results']")
# ActionChains(driver).scroll_to_element(element).perform()

scrollorigin = ScrollOrigin.from_element(element)
print(scrollorigin)
scrolltarget = 3000

e1 = 0

for i in range(80):
    print("i :", i)
    try:
        driver.find_element(By.XPATH, "//span[@class='HlvSq']")
        print("E0.1")
        ActionChains(driver).scroll_from_origin(scrollorigin, 0, 100).perform()
        break
    except:
        try:
            ActionChains(driver).scroll_from_origin(scrollorigin, 0, scrolltarget).perform()
            # scrollorigin = scrollorigin + 3000
            scrolltarget = scrolltarget + 4000
            # print(scrolltarget)
            time.sleep(2)
            e1 = 0
        except:
            time.sleep(2)
            print("E1")
            e1 = e1 + 1
            if e1 > 10:
                break
            else:
                pass
            continue

time.sleep(2)

tiles = driver.find_elements(By.XPATH, "//a[@class='hfpxzc']")
print(len(tiles))

for i in range(len(tiles)):
    try:
        tiles[i].click()
        time.sleep(3)
        # t_data = driver.find_elements(By.CSS_SELECTOR, "div[class='Io6YTe fontBodyMedium kR99db ']")
        try:
            name = driver.find_element(By.CSS_SELECTOR, ".DUwDvf.lfPIob").text
        except:
            name = "except_name"
        try:
            phone = driver.find_element(By.CSS_SELECTOR, '[data-value="Call phone number"]').get_attribute("href")
        except:
            phone = "except_phone"
        print(name + " : " + phone)
        # print(phone.text)
        ws.cell(row=i+1+maxrow, column=1).value = name
        ws.cell(row=i+1+maxrow, column=2).value = phone
        maxrow_n = ws.max_row

        # for i in range(len(t_data)):
        #     ws.cell(row= maxrow_n + 1, column=3 + i).value = t_data[i].text
        wb.save('L_gen.xlsx')

    except:
        print("not clicked", str(i))
        continue

    time.sleep(2)

    name = driver.find_element(By.CSS_SELECTOR, ".DUwDvf.lfPIob").text
    print(name)
    time.sleep(2)

wb.save('L_gen.xlsx')
maxrow = ws.max_row
counts = maxrow - s_max

# grab_data()
if e1 > 10:
    notification.notify("MAP", "Error")
    # push = pb.push_note(url, 'error')
else:
    notification.notify("Succeeded with", str(counts))
    # push = pb.push_note(str(counts), 'Finished')

ws.cell(row= maxrow + 1, column=1).value = "xxxxx" + url
wb.save('L_gen.xlsx')

# //span[@class='UsdlK']
#     driver.find_element(By.XPATH, "//a[@class='hfpxzc']")
