import time
from selenium import webdriver
from selenium.webdriver import ActionChains
from selenium.webdriver.common.actions.wheel_input import ScrollOrigin
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl import worksheet
from pushbullet import Pushbullet
from plyer import notification


API_KEY = "o.QVhTjKwCECLRDhVEUPR8PJNhrIAgxipO"
pb = Pushbullet(API_KEY)

wb = load_workbook('mapinX.xlsx')
ws = wb.active

mx = ws.max_row
print(mx)

driver = webdriver.Chrome()

for i in range(mx):
    if i % 100 == 0:
        # push = pb.push_note(str(i), 'Reach')
        notification.notify("MAP", str(i))
    if ws.cell(row = i+1, column = 2).value == None:
        try:
            driver.get(ws.cell(row=i + 1, column=3).value)
        except:
            print("No Link")
            continue
        time.sleep(1)
        try:
            try:
                name = driver.find_element(By.CSS_SELECTOR, ".DUwDvf.lfPIob").text
                ws.cell(row=i + 1, column=1).value = name
            except:
                ws.cell(row=i + 1, column=1).value = "errorrr"
                pass
            try:
                ph1 = driver.find_element(By.XPATH, "(//div[contains(@class,'rogA2c')])[1]").text
                ph2 = driver.find_element(By.XPATH, "(//div[contains(@class,'rogA2c')])[2]").text
                ph3 = driver.find_element(By.XPATH, "(//div[contains(@class,'rogA2c')])[3]").text
                ph4 = driver.find_element(By.XPATH, "(//div[contains(@class,'rogA2c')])[4]").text
                # ph = driver.find_element(By.CSS_SELECTOR, "a[class='lcr4fd S9kvJb ']").get_attribute('href')
                ws.cell(row=i + 1, column=2).value = "tried"
                ws.cell(row=i + 1, column=4).value = ph1
                ws.cell(row=i + 1, column=5).value = ph2
                ws.cell(row=i + 1, column=6).value = ph3
                ws.cell(row=i + 1, column=7).value = ph4
                # print(name," = ", ph)
                wb.save("mapinX.xlsx")
            except:
                ws.cell(row=i + 1, column=2).value = "errorrr"
                wb.save("mapinX.xlsx")
                pass
            # print(ph)
        except:
            wb.save("mapinX.xlsx")
            continue

    else:
        continue


